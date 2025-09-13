"""
gold_shop.py
Local Desktop Gold Shop app (Tkinter + openpyxl)
- Single-page inputs (rate, customer, carat, weight, making, discount, GST)
- Calculates Gold value, taxable value, CGST/SGST or IGST, Grand Total
- Saves sales to sales.xlsx and generates formatted Excel invoice in invoices/
- Auto-opens invoice after generation
"""

import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
import os
import sys
import subprocess

# Files / folders
SALES_FILE = "sales.xlsx"
INVOICES_DIR = "invoices"

# Ensure files/folders exist
def ensure_files_exist():
    if not os.path.exists(SALES_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales"
        ws.append([
            "DateTime", "Customer", "Phone", "Carat", "Weight(g)",
            "Rate_per_gram", "Gold_Value", "Making", "Discount",
            "Taxable_Value", "GST_Type", "GST_Rate(%)",
            "CGST", "SGST", "IGST", "GST_Total", "Grand_Total", "Invoice_File"
        ])
        wb.save(SALES_FILE)
    if not os.path.exists(INVOICES_DIR):
        os.makedirs(INVOICES_DIR)

# Cross-platform open file
def open_file(path):
    if not os.path.exists(path):
        messagebox.showinfo("Not found", f"File not found: {path}")
        return
    try:
        if sys.platform.startswith('win'):
            os.startfile(os.path.abspath(path))
        elif sys.platform == 'darwin':
            subprocess.call(['open', path])
        else:
            subprocess.call(['xdg-open', path])
    except Exception as e:
        messagebox.showerror("Error", f"Could not open file: {e}")

# Core calculation & persistence
def generate_invoice():
    # Read inputs and validate
    cust = entry_customer.get().strip()
    phone = entry_phone.get().strip()
    rate_s = entry_rate.get().strip()
    carat_s = comb_carat.get().strip()
    weight_s = entry_weight.get().strip()
    making_s = entry_making.get().strip() or "0"
    discount_s = entry_discount.get().strip() or "0"
    gst_rate_s = entry_gst.get().strip() or "0"
    gst_type = comb_gst_type.get().strip()

    # Basic validation
    if not cust:
        messagebox.showerror("Validation", "Enter customer name.")
        return
    if not phone.isdigit() or len(phone) < 6:
        messagebox.showerror("Validation", "Enter valid phone (digits only).")
        return
    try:
        rate = float(rate_s)
        weight = float(weight_s)
        carat = int(carat_s)
        making = float(making_s)
        discount = float(discount_s)
        gst_rate = float(gst_rate_s)
    except ValueError:
        messagebox.showerror("Validation", "Enter valid numeric values for rate/weight/making/discount/GST.")
        return
    if carat <= 0 or carat > 24:
        messagebox.showerror("Validation", "Carat must be between 1 and 24.")
        return
    if weight <= 0:
        messagebox.showerror("Validation", "Weight must be positive.")
        return
    if gst_type not in ("Intra-state (CGST+SGST)", "Inter-state (IGST)"):
        messagebox.showerror("Validation", "Select GST Type.")
        return

    # Calculations
    purity = carat / 24.0
    gold_value = weight * purity * rate           # basic gold value
    taxable = gold_value + making - discount      # taxable value (could add rounding rules)
    gst_total = taxable * (gst_rate / 100.0)
    cgst = sgst = igst = 0.0
    if gst_type == "Intra-state (CGST+SGST)":
        cgst = gst_total / 2.0
        sgst = gst_total / 2.0
    else:
        igst = gst_total

    grand_total = taxable + gst_total

    # Round monetary values sensibly
    gold_value = round(gold_value, 2)
    taxable = round(taxable, 2)
    gst_total = round(gst_total, 2)
    cgst = round(cgst, 2)
    sgst = round(sgst, 2)
    igst = round(igst, 2)
    grand_total = round(grand_total, 2)

    # Save to sales.xlsx
    try:
        wb = openpyxl.load_workbook(SALES_FILE)
        ws = wb.active
        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"), cust, phone, carat, weight,
            rate, gold_value, making, discount, taxable, gst_type, gst_rate,
            cgst if cgst else "", sgst if sgst else "", igst if igst else "", gst_total,
            grand_total, ""  # invoice filename will be updated after creation
        ])
        wb.save(SALES_FILE)
    except Exception as e:
        messagebox.showerror("Error", f"Failed saving sale record: {e}")
        return

    # Create formatted invoice Excel
    try:
        invoice_path = create_formatted_invoice(
            cust, phone, carat, weight, rate, gold_value, making, discount,
            taxable, gst_type, gst_rate, cgst, sgst, igst, gst_total, grand_total
        )
        # Update last row invoice filename in sales.xlsx
        wb = openpyxl.load_workbook(SALES_FILE)
        ws = wb.active
        last_row = ws.max_row
        ws.cell(row=last_row, column=18).value = os.path.basename(invoice_path)
        wb.save(SALES_FILE)

        messagebox.showinfo("Invoice Created", f"Invoice saved:\n{invoice_path}")
        open_file(invoice_path)
    except Exception as e:
        messagebox.showerror("Error", f"Failed creating invoice file: {e}")

# Create invoice Excel with formatting and GST breakup
def create_formatted_invoice(customer, phone, carat, weight, rate, gold_value,
                             making, discount, taxable, gst_type, gst_rate,
                             cgst, sgst, igst, gst_total, grand_total):

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"Invoice_{ts}.xlsx"
    path = os.path.join(INVOICES_DIR, fname)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # Styles
    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Shop header - change to your shop details
    shop_name = "✨ SHREE GOLD & JEWELLERS ✨"
    shop_addr = "123 Main Road, Your City - PIN 400001"
    shop_phone = "Phone: +91-9876543210"
    gstin = "GSTIN: 27ABCDE1234F1Z5"   # Replace with real GSTIN if needed

    ws.merge_cells("A1:D1")
    ws["A1"] = shop_name
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = center

    ws.merge_cells("A2:D2")
    ws["A2"] = shop_addr
    ws["A2"].alignment = center
    ws["A3"] = shop_phone
    ws["C3"] = gstin

    # Invoice meta
    ws["A5"] = "Invoice Date"
    ws["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A6"] = "Customer"
    ws["B6"] = customer
    ws["A7"] = "Phone"
    ws["B7"] = phone

    # Table header
    start_row = 9
    headers = ["Description", "Carat", "Weight (g)", "Rate per g (₹)", "Amount (₹)"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=i, value=h)
        cell.font = bold
        cell.alignment = center
        cell.border = border

    # Row: Gold
    row = start_row + 1
    ws.cell(row=row, column=1, value="Gold Value").border = border
    ws.cell(row=row, column=2, value=carat).border = border
    ws.cell(row=row, column=3, value=weight).border = border
    ws.cell(row=row, column=4, value=rate).border = border
    ws.cell(row=row, column=5, value=gold_value).border = border
    ws.cell(row=row, column=5).alignment = right

    # Making row
    row += 1
    ws.cell(row=row, column=1, value="Making Charges").border = border
    ws.cell(row=row, column=5, value=making).border = border
    ws.cell(row=row, column=5).alignment = right

    # Discount row
    row += 1
    ws.cell(row=row, column=1, value="Discount").border = border
    ws.cell(row=row, column=5, value=-abs(discount)).border = border  # negative for discount
    ws.cell(row=row, column=5).alignment = right

    # Taxable
    row += 1
    ws.cell(row=row, column=4, value="Taxable Value").font = bold
    ws.cell(row=row, column=5, value=taxable).font = bold
    ws.cell(row=row, column=5).alignment = right

    # GST lines
    row += 2
    ws.cell(row=row, column=1, value="GST Type").font = bold
    ws.cell(row=row, column=2, value=gst_type)
    row += 1
    ws.cell(row=row, column=1, value="GST Rate (%)").font = bold
    ws.cell(row=row, column=2, value=gst_rate)

    # Breakup
    row += 2
    if cgst or sgst:
        ws.cell(row=row, column=1, value="CGST").font = bold
        ws.cell(row=row, column=2, value=cgst)
        row += 1
        ws.cell(row=row, column=1, value="SGST").font = bold
        ws.cell(row=row, column=2, value=sgst)
    else:
        ws.cell(row=row, column=1, value="IGST").font = bold
        ws.cell(row=row, column=2, value=igst)

    # GST Total
    row += 2
    ws.cell(row=row, column=4, value="GST Total").font = bold
    ws.cell(row=row, column=5, value=gst_total).font = bold
    ws.cell(row=row, column=5).alignment = right

    # Grand total
    row += 1
    ws.cell(row=row, column=4, value="Grand Total").font = Font(size=12, bold=True)
    ws.cell(row=row, column=5, value=grand_total).font = Font(size=12, bold=True)
    ws.cell(row=row, column=5).alignment = right

    # Footer / declaration
    row += 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="This is a computer generated invoice. No signature required.").alignment = Alignment(horizontal="center")

    # Adjust column widths
    col_widths = {1:28, 2:10, 3:12, 4:16, 5:16}
    for col, w in col_widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    wb.save(path)
    return path

# GUI building
def build_gui():
    ensure_files_exist()
    root = tk.Tk()
    root.title("Gold Shop - GST Ready")
    root.geometry("600x560")
    root.resizable(False, False)
    pad = 6

    frame = ttk.Frame(root, padding=12)
    frame.pack(fill=tk.BOTH, expand=True)

    # Row 0: Rate
    ttk.Label(frame, text="Gold Rate (₹ per gram):").grid(row=0, column=0, sticky="w", pady=pad)
    entry_rate.grid(row=0, column=1, sticky="w")

    # Customer
    ttk.Label(frame, text="Customer Name:").grid(row=1, column=0, sticky="w", pady=pad)
    entry_customer.grid(row=1, column=1, sticky="w")

    ttk.Label(frame, text="Phone:").grid(row=2, column=0, sticky="w", pady=pad)
    entry_phone.grid(row=2, column=1, sticky="w")

    # Carat & weight
    ttk.Label(frame, text="Carat:").grid(row=3, column=0, sticky="w", pady=pad)
    comb_carat.grid(row=3, column=1, sticky="w")

    ttk.Label(frame, text="Weight (grams):").grid(row=4, column=0, sticky="w", pady=pad)
    entry_weight.grid(row=4, column=1, sticky="w")

    # Making & Discount
    ttk.Label(frame, text="Making Charges (₹):").grid(row=5, column=0, sticky="w", pady=pad)
    entry_making.grid(row=5, column=1, sticky="w")

    ttk.Label(frame, text="Discount (₹):").grid(row=6, column=0, sticky="w", pady=pad)
    entry_discount.grid(row=6, column=1, sticky="w")

    # GST controls
    ttk.Label(frame, text="GST Type:").grid(row=7, column=0, sticky="w", pady=pad)
    comb_gst_type.grid(row=7, column=1, sticky="w")
    ttk.Label(frame, text="GST Rate (%):").grid(row=8, column=0, sticky="w", pady=pad)
    entry_gst.grid(row=8, column=1, sticky="w")

    # Buttons
    ttk.Button(frame, text="Generate Invoice (Calculate)", command=generate_invoice).grid(row=9, column=0, columnspan=2, pady=16)
    ttk.Button(frame, text="Open Sales File", command=lambda: open_file(SALES_FILE)).grid(row=10, column=0, pady=6)
    ttk.Button(frame, text="Open Invoices Folder", command=lambda: open_file(INVOICES_DIR)).grid(row=10, column=1, pady=6)

    # Result label
    lbl_result.grid(row=11, column=0, columnspan=2, pady=12)

    # Layout tweaks
    frame.grid_columnconfigure(0, minsize=180)
    frame.grid_columnconfigure(1, minsize=300)

    root.mainloop()

# Widgets
entry_rate = ttk.Entry(width=20)
entry_customer = ttk.Entry(width=30)
entry_phone = ttk.Entry(width=20)
comb_carat = ttk.Combobox(values=[18, 20, 22, 24], width=6)
comb_carat.set(22)
entry_weight = ttk.Entry(width=20)
entry_making = ttk.Entry(width=20)
entry_making.insert(0, "0")
entry_discount = ttk.Entry(width=20)
entry_discount.insert(0, "0")
comb_gst_type = ttk.Combobox(values=["Intra-state (CGST+SGST)", "Inter-state (IGST)"], width=22)
comb_gst_type.set("Intra-state (CGST+SGST)")
entry_gst = ttk.Entry(width=10)
entry_gst.insert(0, "3.0")  # default GST; change as per your shop's rule
lbl_result = ttk.Label(text="", font=("Arial", 14))

if __name__ == "__main__":
    build_gui()
